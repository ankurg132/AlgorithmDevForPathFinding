import openpyxl
from tabulate import tabulate

wb = openpyxl.load_workbook('Project Data.xlsx')

Data = []

for name in wb.sheetnames:
    sheet = wb[name]
    sheet_contents = []
    Cell = lambda r, c: sheet.cell(row = r, column = c).value
    records = Cell(1, 1)
    columns_per_record = Cell(1, 2)
    column = 1
    for record in range(records):
        row = 3
        column = column + columns_per_record
        current_record = []
        while(Cell(row, column - columns_per_record) != 'END'):
            Row = []
            for col in range(column - columns_per_record, column):
                value = Cell(row, col)
                Row.append(value)
            current_record.append(Row)
            row = row + 1
        sheet_contents.append(current_record)
    Data.append([name, sheet_contents])
  
try:
    print("Welcome to flight booking terminal\n1.Domestic flights\n2.International flights")
    choice = int(input("Please enter your choice: "))
    print()
    if choice not in [1, 2]:
        raise Exception("Choice out of range")
except:
    print("You have given invalid input")

if choice == 1:
    num = 0
    print("Welcome to Domestic flight section")
    for record in Data[1][1]:
        num = num + 1
        print('%s. %s'%(num, record[0][0]))
    try:
        countryChoice = int(input("Please choose your country: "))
        print()
        if countryChoice < 1 or countryChoice > num:
            raise Exception("Choice out of Range")
    except:
        print("You have given invalid input")
    print("Departure: ")
    for index in range(1, len(Data[1][1][countryChoice - 1])):
        print('%s. %s'%(index, Data[1][1][countryChoice - 1][index][0]))
    try:
        origin = int(input("Enter choice: "))
        print()
        if origin < 1 or origin > index:
            raise Exception("Choice out of range")
    except:
        print("You have given invalid input")
    num = 0
    for index in range(1, len(Data[1][1][countryChoice - 1])):
        if index == originIndex:
            num = 1
        if index != originIndex:
            print('%s. %s'%(index - num, Data[1][1][countryChoice - 1][index][0]))
    try:
        destinationIndex = int(input("Enter choice: "))
        print()
        if originIndex > destinationIndex:
            destination = Data[1][1][countryChoice - 1][destinationIndex][0]
        else:
            destination = Data[1][1][countryChoice - 1][destinationIndex + 1][0]
        if destinationIndex < 1 or destinationIndex > index:
            raise Exception("Choice out of range")
    except:
        print("You have given invalid input")
    flag = False
    for pair in Data[0][1][0]:
        if pair[0] == origin:
            originCode = pair[1]
            if flag:
                break
            else:
                flag = True
        if pair[0] == destination:
            destinationCode = pair[1]
            if flag:
                break
            else:
                flag = True
    print("The origin of your journer: %s\nYour destination: %s"%(origin, destination))
    print()
    flightDetails = []
    index = 0
    for sheet in Data:
        if sheet[0] == 'FLIGHT_PLANS_'+country:
            for cityRecord in sheet[1]:
                if cityRecord[0][0] == originCode:
                    for row in cityRecord:
                        if row[1] == destinationCode:
                            index = index + 1
                            Row = [index] + row
                            flightDetails.append(Row)
    print(tabulate(flightDetails, headers = ['SNo.', 'From', 'To', 'Duration', 'Flight Details', 'Cost']))
else:
    print("This section is under development")
